"""
@project: Learn_DP
@author: linjo
@file: Products.py
@date: 2019-10-19
@time: 12:02 PM

"""

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from pandas import DataFrame


# Get Items
def get_items(urlPath):
    # Request
    res = requests.get(urlPath, headers)
    soup = BeautifulSoup(res.text, 'html.parser')

    # find all items: grid-tile
    grid_tiles = soup.find_all('div', {'class': 'grid-tile'})
    for grid_tile in grid_tiles:
        try:
            # item-tile
            item_tile = grid_tile.find('div', {'class': 'product-tile'}).get('data-cgid')
            tile_list.append(item_tile)

            # item-name
            item_name = grid_tile.find('a', {'class': 'name-link'}).get('title')  # == .string
            name_list.append(item_name)

            # item-price & items link
            item_price = grid_tile.find('span', {'class': 'actual-price'}).get_text()  # == .string
            if ',' in item_price:
                temp = float(item_price.replace(',', '')[1:]) * 1.13
                price_list.append(temp)
            else:
                temp2 = float(item_price[1:]) * 1.13
                price_list.append(temp2)

            # item-color
            item_colors = grid_tile.find('ul', {'class': 'swatch-list'})
            colors = []
            for c in item_colors.find_all('a', {'class': 'swatch'}):
                colors.append(c.get('title'))
            color_list.append(str(colors))

            # item-attributes
            item_attributes = grid_tile.find('div', {'class': 'plp-custom-attributes'})
            attributes = []
            for i in item_attributes.find_all('span', {'class': 'plp-attribute'}):
                attributes.append(i.get_text())
            attributes_list.append(str(attributes))

            # item-size: now-avaliable & item-link
            # Get item details page
            item_details = "https://www.canadagoose.com" \
                           + grid_tile.find('div', {'class': 'product-image'}).find('a',
                                                                                    {'class': 'thumb-link'}).get(
                'href')

            # if the item is not avaliable any more
            try:
                res = requests.get(item_details, headers)
                soup_details = BeautifulSoup(res.text, 'html.parser')
                # size list
                size_lists = soup_details.find('div', {'class': 'size-list'})
                sizes = []
                for size_details in size_lists.find_all('div'):
                    if len(size_details.get('class')) == 1:
                        sizes.append(size_details.find('a').get('data-sizeval'))
                size_list.append(str(sizes))
                link_list.append(item_details)
            except Exception as e:
                print('item is not avaliable ! ')
                size_list.append("null")
                link_list.append("Item is not avaliable")

        except AttributeError as e:
            pass

    # Check if it has more items, then call function again
    try:
        data_grid_url = soup.find('div', {'class': 'infinite-scroll-placeholder'}).get('data-grid-url')
        print(data_grid_url)
        if data_grid_url != "":
            print("continue")
            get_items(data_grid_url)
    except Exception as e:
        print("No More Items!")
        pass


if __name__ == '__main__':
    # Request url
    headers = {
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.120 Safari/537.36'
    }

    # params
    params2 = []

    for i in params2:
        # List
        tile_list = []
        name_list = []
        price_list = []
        size_list = []
        color_list = []
        attributes_list = []
        link_list = []

        # call: get_items
        temp_link = 'https://www.canadagoose.com/ca/en/shop' + i
        get_items(temp_link)
        print(tile_list)

        data = {
            'Title': tile_list,
            'Name': name_list,
            'Price + Tax': price_list,
            'Size': size_list,
            'Color': color_list,
            'Attributes': attributes_list,
            'Link': link_list
        }

        ws = load_workbook('Goose.xlsx')
        if tile_list[0] in ws.get_sheet_names():
            temp = ws.get_sheet_by_name(tile_list[0])
            ws.remove_sheet(temp)
            ws.save('Goose.xlsx')

        book = load_workbook('Goose.xlsx')
        df = DataFrame(data)
        with pd.ExcelWriter('Goose.xlsx', engine='openpyxl', mode='a') as writer:
            df.to_excel(writer, sheet_name=tile_list[0])
